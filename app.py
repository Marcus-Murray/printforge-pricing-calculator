"""
PrintForge Pricing Calculator - Flask Web Application
Clean web-based interface with Ant Design styling
"""

from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tempfile

app = Flask(__name__)
app.config['SECRET_KEY'] = 'printforge-pricing-2026'

# Ensure directories exist
UPLOAD_FOLDER = Path('uploads')
UPLOAD_FOLDER.mkdir(exist_ok=True)

@app.route('/')
def index():
    """Main application page"""
    return render_template('index.html')

@app.route('/calculate', methods=['POST'])
def calculate():
    """Calculate pricing based on input data"""
    try:
        data = request.json
        
        # Extract basic info
        filament_cost = float(data.get('filament_cost', 40.0))
        filament_required = float(data.get('filament_required', 0.0))
        print_time = float(data.get('print_time', 0.0))
        labor_time = float(data.get('labor_time', 0))
        
        # Extract hardware items
        hardware_items = data.get('hardware_items', [])
        hardware_total = sum(item['quantity'] * item['unit_cost'] 
                           for item in hardware_items)
        
        # Extract packaging items
        packaging_items = data.get('packaging_items', [])
        packaging_total = sum(item['quantity'] * item['unit_cost'] 
                            for item in packaging_items)
        shipping_cost = float(data.get('shipping_cost', 0.0))
        
        # Extract advanced settings
        printer_cost = float(data.get('printer_cost', 1000.0))
        upfront_cost = float(data.get('upfront_cost', 0.0))
        annual_maintenance = float(data.get('annual_maintenance', 75.0))
        printer_life = float(data.get('printer_life', 3.0))
        average_uptime = float(data.get('average_uptime', 50.0))
        
        power_consumption = float(data.get('power_consumption', 250.0))
        electricity_rate = float(data.get('electricity_rate', 0.30))
        electricity_daily = float(data.get('electricity_daily', 1.50))
        
        efficiency_factor = float(data.get('efficiency_factor', 1.1))
        labor_rate = float(data.get('labor_rate', 20.0))
        
        # Calculate material cost
        material_cost = (filament_required / 1000) * filament_cost * efficiency_factor
        
        # Calculate labor cost
        labor_cost = (labor_time / 60) * labor_rate
        
        # Calculate machine depreciation cost
        total_printer_cost = printer_cost + upfront_cost
        annual_cost = annual_maintenance
        lifetime_cost = total_printer_cost + (annual_cost * printer_life)
        
        hours_per_year = 365.25 * 24
        uptime_fraction = average_uptime / 100
        total_uptime_hours = hours_per_year * printer_life * uptime_fraction
        
        cost_per_hour = lifetime_cost / total_uptime_hours if total_uptime_hours > 0 else 0
        machine_depreciation = print_time * cost_per_hour
        
        # Calculate electricity cost
        power_kw = power_consumption / 1000
        electricity_usage = power_kw * print_time * electricity_rate
        electricity_fixed = (print_time / 24) * electricity_daily
        electricity_cost = electricity_usage + electricity_fixed
        
        machine_cost_total = machine_depreciation + electricity_cost
        
        # Calculate packaging cost
        total_packaging = packaging_total + shipping_cost
        
        # Calculate total landed cost
        total_cost = material_cost + labor_cost + machine_cost_total + total_packaging
        
        # Calculate pricing at different margins
        def calc_price(margin):
            return total_cost / (1 - margin / 100) if margin < 100 else 0
        
        price_50 = calc_price(50)
        price_60 = calc_price(60)
        price_70 = calc_price(70)
        custom_margin = float(data.get('custom_margin', 75))
        price_custom = calc_price(custom_margin)
        
        return jsonify({
            'success': True,
            'material_cost': round(material_cost, 2),
            'labor_cost': round(labor_cost, 2),
            'machine_depreciation': round(machine_depreciation, 2),
            'electricity_cost': round(electricity_cost, 2),
            'machine_cost_total': round(machine_cost_total, 2),
            'packaging_cost': round(total_packaging, 2),
            'total_cost': round(total_cost, 2),
            'price_50': round(price_50, 2),
            'price_60': round(price_60, 2),
            'price_70': round(price_70, 2),
            'price_custom': round(price_custom, 2),
            'custom_margin': custom_margin,
            'cost_per_hour': round(cost_per_hour, 4)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/save-config', methods=['POST'])
def save_config():
    """Save configuration to JSON file"""
    try:
        data = request.json
        filename = data.get('filename', 'config.json')
        config_data = data.get('config', {})
        
        # Add metadata
        config_data['version'] = '1.0'
        config_data['saved_date'] = datetime.now().isoformat()
        
        filepath = UPLOAD_FOLDER / filename
        with open(filepath, 'w') as f:
            json.dump(config_data, f, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'Configuration saved as {filename}'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/load-config', methods=['POST'])
def load_config():
    """Load configuration from JSON file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        config_data = json.load(file)
        
        return jsonify({
            'success': True,
            'config': config_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/export-excel', methods=['POST'])
def export_excel():
    """Export pricing data to Excel"""
    try:
        data = request.json
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Pricing Breakdown"
        
        # Define styles
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
        section_font = Font(name='Arial', size=12, bold=True)
        section_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        label_font = Font(name='Arial', size=10, bold=True)
        value_font = Font(name='Arial', size=10)
        currency_font = Font(name='Arial', size=10, bold=True, color='FF6B35')
        total_font = Font(name='Arial', size=14, bold=True, color='FF6B35')
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "PRINTFORGE PRICING BREAKDOWN"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 2
        
        # Product Information
        ws[f'A{row}'] = "PRODUCT INFORMATION"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        product_info = [
            ("Part Name:", data.get('part_name', 'New Part')),
            ("Revision:", data.get('revision', 'V1')),
            ("Prepared By:", data.get('prepared_by', 'Marcus')),
            ("Date:", datetime.now().strftime("%Y-%m-%d")),
            ("Material:", data.get('material_type', 'ABS')),
        ]
        
        for label, value in product_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = value_font
            row += 1
        row += 1
        
        # Material & Print Settings
        ws[f'A{row}'] = "MATERIAL & PRINT SETTINGS"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        print_settings = [
            ("Filament Cost:", f"${data.get('filament_cost', 40):.2f} /kg"),
            ("Filament Required:", f"{data.get('filament_required', 0):.2f} g"),
            ("Print Time:", f"{data.get('print_time', 0):.2f} hours"),
            ("Labor Time:", f"{data.get('labor_time', 0)} minutes"),
        ]
        
        for label, value in print_settings:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = value_font
            row += 1
        row += 1
        
        # Hardware Components
        hardware_items = data.get('hardware_items', [])
        if hardware_items:
            ws[f'A{row}'] = "HARDWARE COMPONENTS"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # Headers
            headers = ["Name", "Quantity", "Unit Cost", "Total"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = label_font
                cell.border = border
            row += 1
            
            for item in hardware_items:
                ws[f'A{row}'] = item.get('name', '')
                ws[f'B{row}'] = item.get('quantity', 0)
                ws[f'C{row}'] = f"${item.get('unit_cost', 0):.2f}"
                total = item.get('quantity', 0) * item.get('unit_cost', 0)
                ws[f'D{row}'] = f"${total:.2f}"
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].font = value_font
                    ws[f'{col}{row}'].border = border
                row += 1
            row += 1
        
        # Packaging Materials
        packaging_items = data.get('packaging_items', [])
        shipping_cost = data.get('shipping_cost', 0)
        if packaging_items or shipping_cost > 0:
            ws[f'A{row}'] = "PACKAGING & SHIPPING"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # Headers
            headers = ["Name", "Quantity", "Unit Cost", "Total"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = label_font
                cell.border = border
            row += 1
            
            for item in packaging_items:
                ws[f'A{row}'] = item.get('name', '')
                ws[f'B{row}'] = item.get('quantity', 0)
                ws[f'C{row}'] = f"${item.get('unit_cost', 0):.2f}"
                total = item.get('quantity', 0) * item.get('unit_cost', 0)
                ws[f'D{row}'] = f"${total:.2f}"
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].font = value_font
                    ws[f'{col}{row}'].border = border
                row += 1
            
            if shipping_cost > 0:
                ws[f'A{row}'] = "Shipping"
                ws[f'B{row}'] = 1
                ws[f'C{row}'] = f"${shipping_cost:.2f}"
                ws[f'D{row}'] = f"${shipping_cost:.2f}"
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].font = value_font
                    ws[f'{col}{row}'].border = border
                row += 1
            row += 1
        
        # Cost Breakdown
        results = data.get('results', {})
        ws[f'A{row}'] = "COST BREAKDOWN"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        cost_breakdown = [
            ("Materials Cost:", f"${results.get('material_cost', 0):.2f}"),
            ("Labor Cost:", f"${results.get('labor_cost', 0):.2f}"),
            ("Machine Cost:", f"${results.get('machine_cost_total', 0):.2f}"),
            ("Packaging & Shipping:", f"${results.get('packaging_cost', 0):.2f}"),
        ]
        
        for label, value in cost_breakdown:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = currency_font
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        # Total Landed Cost
        row += 1
        ws[f'A{row}'] = "TOTAL LANDED COST:"
        ws[f'A{row}'].font = total_font
        ws[f'C{row}'] = f"${results.get('total_cost', 0):.2f}"
        ws[f'C{row}'].font = total_font
        ws[f'C{row}'].alignment = Alignment(horizontal='right')
        row += 2
        
        # Suggested Pricing
        ws[f'A{row}'] = "SUGGESTED PRICING"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        pricing = [
            ("50% Margin:", f"${results.get('price_50', 0):.2f}"),
            ("60% Margin:", f"${results.get('price_60', 0):.2f}"),
            ("70% Margin:", f"${results.get('price_70', 0):.2f}"),
            (f"{results.get('custom_margin', 75)}% Margin (Custom):", f"${results.get('price_custom', 0):.2f}"),
        ]
        
        for label, value in pricing:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = currency_font
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Save to temporary file
        part_name = data.get('part_name', 'New_Part').replace(' ', '_')
        filename = f"{part_name}_Pricing.xlsx"
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

if __name__ == '__main__':
    print("\n" + "="*60)
    print("PrintForge Pricing Calculator")
    print("="*60)
    print("\nOpening in your browser...")
    print("URL: http://localhost:5000")
    print("\nPress Ctrl+C to stop the server")
    print("="*60 + "\n")
    
    # Auto-open browser
    import webbrowser
    from threading import Timer
    Timer(1.5, lambda: webbrowser.open('http://localhost:5000')).start()
    
    app.run(debug=True, port=5000)
